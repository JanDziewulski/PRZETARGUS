# Pckg install
from RegonAPI import RegonAPI
from RegonAPI.exceptions import ApiAuthenticationError
import re
from pdfminer import high_level
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.converter import XMLConverter, HTMLConverter, TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfparser import PDFParser
import io
import os
import time
from openpyxl import load_workbook
# import frontend as fr
import statistics
from statistics import mode

class PDFsearcher:

    def pdf_read(self, file):
        # pdf_data = fr.browseFiles()
        fp = open(file, 'rb')
        rsrcmgr = PDFResourceManager()
        retstr = io.StringIO()
        print(type(retstr))
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        page_no = 0
        for pageNumber, page in enumerate(PDFPage.get_pages(fp)):
            if pageNumber == page_no:
                interpreter.process_page(page)

                data = retstr.getvalue()

            page_no += 1
        return data

    def niptofront(self, nip):
        # Available reports
        REPORTS = ["BIR11OsFizycznaDaneOgolne"]
        c_nip = re.sub(r'[^\d]', '', nip)
        # TEST_API_KEY = 'b7cbd57bd28f4f4c99fe' - klucz produkcyjny
        TEST_API_KEY = "abcde12345abcde12345"
        CD_PROJEKT_NIP = f'{c_nip}'

        # Authentication
        api = RegonAPI(bir_version="bir1.1", is_production=False)
        try:
            api.authenticate(key=TEST_API_KEY)
        except ApiAuthenticationError as e:
            print("[-]", e)
            exit(0)
        except Exception as e:
            raise

        # Search by NIP
        result = api.searchData(nip=CD_PROJEKT_NIP)

        ulica = result[0].get('Ulica')
        numer = result[0].get('NrNieruchomosci')
        miejscowosc = result[0].get('Miejscowosc')
        kod = result[0].get('KodPocztowy')
        nazwa = result[0].get('Nazwa')
        adres = f'{ulica} {numer}\n {miejscowosc} {kod}'
        return adres, nazwa

#wyszukiwanie danych do okna dialogowego
    def pdf_nip_info(self, file):
        nip = re.search(r'NIP(\W+\d{10})|NIP(\W+PL\d{10})|NIP(\W+PL\d{3}-\d{2,3}-\d{2}-\d{2,3})|NIP(\W+\d{3}-\d{2,3}-\d{2}-\d{2,3})', file)
        return nip.group()

    def pdf_mail_info(self, file):
        mail = re.search(r'(\w+([@])\w+.\w+)|(\w+.\w+([@])\w+.\w+)', file)
        return mail.group()


    def text_recon(self, text, text_to_recognize):
        """Function that checks the presence of text in a pdf file.

            Args:
                text (str): Parameter imported from .xlsx file.

            Returns:
                bool: The return value. True for success, False otherwise.

            """

        """Loading and cleaning pdf file"""
        # local_pdf_filename = "ogloszenie_12215.pdf"  # path to file
        # pages = [7]  # page number

        # extracted_text = high_level.extract_text(local_pdf_filename, "", pages)
        text = text.lower()
        extracted_text_clean = re.sub('[^\s\d\w]', '', text)  # text clean
        # print(extracted_text_clean)

        """Reading and converting searched text"""
        reg_rext = text_to_recognize
        reg_rext_cln = re.sub('[^\s\d\w]', '', reg_rext)  # cleaned text
        text_to_list = reg_rext_cln.lower().split(' ')
        # print(text_to_list)

        """List with item paris"""
        if len(text_to_list) % 2 == 0:
            pair_list = [text_to_list[i] + ' ' + text_to_list[i + 1] for i in range(0, len(text_to_list), 2)]
        else:
            pair_list = [text_to_list[i] + ' ' + text_to_list[i + 1] for i in range(0, len(text_to_list) - 1, 2)]

        # print(pair_list)

        count = 0
        for i in range(len(pair_list)):
            founded_item = re.search(pair_list[i], extracted_text_clean)
            try:
                if founded_item.group() == pair_list[i]:
                    count += 1
            except BaseException:
                pass

                # print('Nie znaleziono tego słowa')

        c_ratio = count / len(pair_list)
        if 1 >= c_ratio >= 0.8:
            return True
            # print('Rozpoznanie z pełnym Sukcesem')
        elif 0.8 > c_ratio >= 0.5:
            print('Sprawdź jeszcze ręcznie ')
            return False
        else:
            return False
            print('Współczynnik rozpoznania {}'.format(round(c_ratio, 2)))

    def text_from_excel(self):

        """
        emp_list[0][0] - opis funkcjonalności, np. Wbudowana baza uchwytów, imadła, uchwyty 3 szczękowe, łapy dociskowe (niedopuszczalne jest dostarczenie bazy osobno poza systemem),
        emp_list[0][1] - nazwa programu, np. Edgecam
        emp_list[0][2] - informacje dodatkowe, np. Nie spełniamy wymagań

        :return:
        """
        global iter_num
        workbook = load_workbook(filename="Wymagania przetargowe.xlsx")
        sheet = workbook.active
        iter_num = sheet.max_row
        apn_list = []
        for value in sheet.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
            apn_list.append(value)
            # print(value)
            # text_recon(value)
        return apn_list, iter_num
        # print(apn_list[1][0])




# text_recon()
if __name__ == '__main__':
    # imput_data= pfd_read('ogloszenie_12215.pdf')
    c = PDFsearcher()

    imput_data= c.pdf_read('ogloszenie_12215.pdf')
    # text_recon(imput_data)
    # print(imput_data)
    # text_from_excel()
    excel_list = c.text_from_excel()
    print(c.pdf_info(imput_data))
    program_list = []
    for i in range(iter_num - 1):
        val = c.text_recon(imput_data, excel_list[i][0])
        if val is True:
            program_list.append(excel_list[i][1])
            print(excel_list[i][1:])
        else:
            print('Nie rozpoznano!')
    print(mode(program_list))

